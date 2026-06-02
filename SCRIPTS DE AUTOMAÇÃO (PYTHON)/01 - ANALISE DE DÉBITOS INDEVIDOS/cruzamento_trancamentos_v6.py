# ============================================================
#  UNIFECAF — CRUZAMENTO DE TRANCAMENTOS x DÉBITOS INDEVIDOS
#  Financeiro — Cobrança | Uso Interno
#  Versão 1.5 | Junho/2026
# ============================================================
#
#  COMO USAR:
#  1. Coloque este arquivo em qualquer pasta do seu computador
#  2. Clique duas vezes nele para abrir
#  3. Informe o caminho dos dois arquivos quando solicitado
#  4. Aguarde — a planilha de resultado será salva na mesma
#     pasta dos arquivos de entrada
#
# ============================================================

import os
import sys
import warnings
warnings.filterwarnings('ignore')

# ── Verificar bibliotecas ────────────────────────────────────
def verificar_bibliotecas():
    faltando = []
    for lib in ['pandas', 'openpyxl', 'xlrd']:
        try:
            __import__(lib)
        except ImportError:
            faltando.append(lib)
    if faltando:
        print("\n❌ ERRO: As seguintes bibliotecas não estão instaladas:")
        for lib in faltando:
            print(f"   - {lib}")
        print("\nAbra o terminal (cmd) e execute:")
        print("   pip install pandas openpyxl xlrd")
        input("\nPressione Enter para fechar...")
        sys.exit(1)

verificar_bibliotecas()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

# ════════════════════════════════════════════════════════════
# CONFIGURAÇÕES VISUAIS
# ════════════════════════════════════════════════════════════
AZUL_ESC  = '002855'; AZUL_MED  = '003D7A'; AZUL      = '0057A8'
AZUL_CLR  = 'E8F1FB'; AZUL_MID  = 'CCE0F5'; VERDE     = '2EAD6A'
VERDE_CLR = 'E4F7EE'; VERDE_ESC = '1B7A3E'; VERM      = 'C0392B'
VERM_CLR  = 'FDECEA'; GOLD      = 'F5A623'; GOLD_CLR  = 'FEF7E8'
CINZA_CLR = 'F4F6F9'; BRANCO    = 'FFFFFF'

def fill(h):
    return PatternFill('solid', fgColor=h)

def font(h, sz=10, bold=False, italic=False):
    return Font(name='Calibri', size=sz, color=h, bold=bold, italic=italic)

def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def brd(color='CCE0F5'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def brd_bottom(color='E0E8F4'):
    return Border(bottom=Side(style='thin', color=color))

# ════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ════════════════════════════════════════════════════════════
def extrair_plano(txt):
    t = str(txt).upper() if not pd.isna(txt) else ''
    if '[ANT]' in t:
        return 'ANT'
    if '[PST]' in t:
        return 'PST'
    return 'OUTRO'

def classificar(plano, data_solic, vencimento, foi_pago):
    """
    Classifica cada parcela como DEVIDA ou INDEVIDA com base no plano
    de pagamento e na data de solicitação do trancamento.

    Lógica de negócio:
    ─────────────────────────────────────────────────────────────────
    PLANO ANT (o aluno paga pelo serviço que AINDA VAI USAR):
      • Solicitação até o dia 07 do mês:
          - Meses ANTERIORES ao mês da solicitação  → DEVIDA
            (o aluno já utilizou o serviço nesses meses)
          - Mês da solicitação e meses SEGUINTES    → INDEVIDA
            (o aluno não vai mais utilizar o serviço)

      • Solicitação a partir do dia 11 do mês:
          - Mês da solicitação e meses ANTERIORES   → DEVIDA
            (ultrapassou o prazo; o compromisso com o mês corrente
             já estava assumido, assim como os anteriores)
          - Meses SEGUINTES ao mês da solicitação   → INDEVIDA
            (serviço não será prestado nos meses seguintes)

    ─────────────────────────────────────────────────────────────────
    PLANO PST (o aluno paga pelo serviço que JÁ UTILIZOU):
      • Solicitação até o dia 07 do mês:
          - Mês da solicitação e meses ANTERIORES   → DEVIDA
            (vencimento do mês corrente refere-se ao mês anterior
             estudado — sempre devido)
          - Meses SEGUINTES ao mês da solicitação   → INDEVIDA
            (não terá acesso ao serviço nesses meses)

      • Solicitação a partir do dia 11 do mês:
          - Mês da solicitação e meses ANTERIORES   → DEVIDA
          - Mês SEGUINTE ao mês da solicitação      → DEVIDA
            (solicitou após o dia 07, portanto já teve acesso
             suficiente no mês corrente — mês seguinte também é devido)
          - Demais meses seguintes                  → INDEVIDA
    ─────────────────────────────────────────────────────────────────
    """
    if pd.isna(data_solic):
        return ('SEM TRANCAMENTO NA BASE',
                'Aluno não localizado na base de requerimentos deferidos')

    dia   = data_solic.day
    ok10  = dia <= 7          # True = solicitou até o dia 07
    d_str = data_solic.strftime('%d/%m/%Y')

    if plano == 'ANT':
        if pd.isna(vencimento):
            return ('VERIFICAR', 'ANT – sem data de vencimento para análise')

        ms = data_solic.to_period('M')   # mês/ano da solicitação
        mv = vencimento.to_period('M')   # mês/ano do vencimento da parcela

        if ok10:
            # ── ANT · Solicitação ATÉ o dia 07 ──────────────────────
            # Meses anteriores ao mês da solicitação → DEVIDA
            # Mês corrente e seguintes               → INDEVIDA
            if mv < ms:
                return ('DEVIDA',
                        f'ANT – solicitação até dia 07 ({d_str}); '
                        f'parcela de mês anterior à solicitação – serviço já utilizado')
            else:
                return ('INDEVIDA',
                        f'ANT – solicitação até dia 07 ({d_str}); '
                        f'mês corrente e seguintes indevidos – serviço não será prestado')
        else:
            # ── ANT · Solicitação A PARTIR do dia 11 ────────────────
            # Mês da solicitação e meses anteriores  → DEVIDA
            # Meses seguintes ao mês da solicitação  → INDEVIDA
            if mv <= ms:
                return ('DEVIDA',
                        f'ANT – solicitação após dia 07 ({d_str}); '
                        f'mês da solicitação e anteriores são devidos – prazo de corte ultrapassado')
            else:
                return ('INDEVIDA',
                        f'ANT – solicitação após dia 07 ({d_str}); '
                        f'mês seguinte à solicitação indevido – serviço não será prestado')

    elif plano == 'PST':
        if pd.isna(vencimento):
            return ('VERIFICAR', 'PST – sem data de vencimento para análise')

        ms = data_solic.to_period('M')   # mês/ano da solicitação
        mv = vencimento.to_period('M')   # mês/ano do vencimento da parcela
        diff = (mv - ms).n               # diferença em meses (0 = mesmo mês, 1 = próximo mês, etc.)

        if ok10:
            # ── PST · Solicitação ATÉ o dia 07 ──────────────────────
            # Mês da solicitação e meses anteriores  → DEVIDA
            #   (vencimento do mês corrente refere-se ao mês anterior estudado)
            # Meses seguintes                        → INDEVIDA
            if mv <= ms:
                return ('DEVIDA',
                        f'PST – solicitação até dia 07 ({d_str}); '
                        f'mês corrente e anteriores são devidos – vencimento refere-se ao mês anterior estudado')
            else:
                return ('INDEVIDA',
                        f'PST – solicitação até dia 07 ({d_str}); '
                        f'meses seguintes indevidos – aluno não terá acesso ao serviço')
        else:
            # ── PST · Solicitação A PARTIR do dia 11 ────────────────
            # Mês da solicitação e anteriores        → DEVIDA
            # Mês imediatamente seguinte             → DEVIDA (teve acesso após dia 07)
            # Demais meses seguintes                 → INDEVIDA
            if mv <= ms:
                return ('DEVIDA',
                        f'PST – solicitação após dia 07 ({d_str}); '
                        f'mês corrente e anteriores são devidos – vencimento refere-se ao mês anterior estudado')
            elif diff == 1:
                return ('DEVIDA',
                        f'PST – solicitação após dia 07 ({d_str}); '
                        f'próximo mês também devido – aluno teve acesso ao serviço após o dia 07')
            else:
                return ('INDEVIDA',
                        f'PST – solicitação após dia 07 ({d_str}); '
                        f'meses seguintes ao próximo indevidos – serviço não será prestado')

    return ('VERIFICAR MANUALMENTE',
            'Plano não identificado (OUTRO) – verificar manualmente')

# ════════════════════════════════════════════════════════════
# ESTILIZAÇÃO DE PLANILHA
# ════════════════════════════════════════════════════════════
def estilo_cabecalho(ws, row, n_cols, bg=AZUL_ESC):
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = fill(bg)
        c.font = font(BRANCO, 10, True)
        c.alignment = aln('center')
        c.border = brd()
    ws.row_dimensions[row].height = 22

def estilo_linha(ws, row, n_cols, bg=BRANCO):
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = fill(bg)
        c.font = font('1E1E1E', 9)
        c.alignment = aln('left')
        c.border = brd_bottom()
    ws.row_dimensions[row].height = 16

def escrever_df(ws, df, start_row=2, header_bg=AZUL_ESC,
                colunas_centro=None, colunas_moeda=None,
                colunas_wrap=None):
    colunas_centro = colunas_centro or []
    colunas_moeda  = colunas_moeda  or []
    colunas_wrap   = colunas_wrap   or []

    # Cabeçalho
    estilo_cabecalho(ws, start_row, len(df.columns), bg=header_bg)
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci).value = col

    # Dados
    COR_CLASS = {
        'INDEVIDA':               (VERM_CLR,  VERM),
        'DEVIDA':                 (VERDE_CLR, VERDE_ESC),
        'MANTER (PAGO)':          (GOLD_CLR,  '8a5e00'),
        'VERIFICAR MANUALMENTE':  (CINZA_CLR, '555555'),
        'SEM TRANCAMENTO NA BASE':('F5F5F5',  '888888'),
        'VERIFICAR':              (GOLD_CLR,  '8a5e00'),
    }

    last_ri = start_row
    for ri, (_, row_data) in enumerate(df.iterrows(), start_row + 1):
        bg_row = CINZA_CLR if (ri - start_row) % 2 == 0 else BRANCO
        ws.row_dimensions[ri].height = 16

        for ci, (col, val) in enumerate(zip(df.columns, row_data), 1):
            # Garantir que a célula não é mesclada antes de escrever
            c = ws.cell(row=ri, column=ci)
            from openpyxl.cell.cell import MergedCell
            if isinstance(c, MergedCell):
                continue

            # Estilo base
            c.fill = fill(bg_row)
            c.font = font('1E1E1E', 9)
            c.alignment = aln('left')
            c.border = brd_bottom()

            # Valor
            if pd.isna(val) if not isinstance(val, str) else False:
                c.value = ''
            elif col in colunas_moeda:
                try:
                    c.value = float(val)
                    c.number_format = 'R$ #,##0.00'
                except Exception:
                    c.value = val
            else:
                c.value = val if val is not None else ''

            # Alinhamento específico
            if col in colunas_centro:
                c.alignment = aln('center')
            elif col in colunas_wrap:
                c.alignment = aln('left', wrap=True)
                ws.row_dimensions[ri].height = 30

            # Cor especial para CLASSIFICAÇÃO
            if col == 'CLASSIFICAÇÃO':
                bg_c, fg_c = COR_CLASS.get(str(val), (BRANCO, '1E1E1E'))
                c.fill = fill(bg_c)
                c.font = font(fg_c, 9, True)
                c.alignment = aln('center')

        last_ri = ri

    return last_ri  # última linha usada

# ════════════════════════════════════════════════════════════
# INTERFACE COM O USUÁRIO
# ════════════════════════════════════════════════════════════
def limpar():
    os.system('cls')

def cabecalho():
    print("=" * 60)
    print("  UNIFECAF — CRUZAMENTO TRANCAMENTOS x DÉBITOS INDEVIDOS")
    print("  Financeiro — Cobrança  |  Uso Interno  |  V1.5")
    print("=" * 60)
    print()

def pedir_arquivo(mensagem, extensoes_validas=('.xlsx', '.xls')):
    while True:
        print(mensagem)
        print("  (Dica: copie o caminho completo do arquivo, ex: C:\\Users\\...)")
        caminho = input("  Caminho: ").strip().strip('"').strip("'")
        if not caminho:
            print("  ❌ Caminho não pode ser vazio. Tente novamente.\n")
            continue
        if not os.path.exists(caminho):
            print(f"  ❌ Arquivo não encontrado: {caminho}")
            print("  Verifique o caminho e tente novamente.\n")
            continue
        ext = os.path.splitext(caminho)[1].lower()
        if ext not in extensoes_validas:
            print(f"  ❌ Formato inválido. Use: {', '.join(extensoes_validas)}\n")
            continue
        print(f"  ✅ Arquivo encontrado!\n")
        return caminho

# ════════════════════════════════════════════════════════════
# PROCESSAMENTO PRINCIPAL
# ════════════════════════════════════════════════════════════
def processar(caminho_auditoria, caminho_requerimentos):

    # ── Carregar bases ───────────────────────────────────────
    print("  [1/6] Carregando base de auditoria financeira...")
    df_audit = pd.read_excel(caminho_auditoria, dtype={'Ra Aluno': str})
    print(f"        → {len(df_audit):,} registros carregados")

    print("  [2/6] Carregando base de requerimentos...")
    df_req = pd.read_excel(caminho_requerimentos, dtype={'Ra do Aluno': str})
    print(f"        → {len(df_req):,} registros carregados")

    # ── Preparar auditoria ───────────────────────────────────
    print("  [3/6] Preparando dados...")
    df_audit['Ra Aluno']   = df_audit['Ra Aluno'].astype(str).str.strip()
    df_audit['Vencimento'] = pd.to_datetime(df_audit['Vencimento'],
                                             dayfirst=True, errors='coerce')
    df_audit['Data Pag']   = pd.to_datetime(df_audit['Data Pag'],
                                             dayfirst=True, errors='coerce')
    df_audit['foi_pago']   = (df_audit['Data Pag'].notna() &
                               (df_audit['Data Pag'].dt.year > 1901))
    df_audit['tipo_plano']  = df_audit['Plano Pagamento'].apply(extrair_plano)
    df_audit['Turma_norm']  = df_audit['Turma'].astype(str).str.strip().str.upper()

    # ── Preparar requerimentos ───────────────────────────────
    df_req['Ra do Aluno'] = df_req['Ra do Aluno'].astype(str).str.strip()
    df_req['Data da solicitação'] = pd.to_datetime(
        df_req['Data da solicitação'], dayfirst=True, errors='coerce')

    df_tranc = df_req[
        df_req['Nome do Requerimento'].str.upper()
            .str.contains('TRANCAMENTO', na=False) &
        df_req['Situação do Requerimento'].str.upper()
            .str.contains('DEFERIDO', na=False)
    ].copy()

    n_tranc = df_tranc['Ra do Aluno'].nunique()
    print(f"        → {n_tranc:,} alunos com trancamento deferido")

    # Manter apenas o trancamento mais recente por aluno
    # Deduplicar por RA + Turma (mantém um trancamento por curso por aluno)
    df_tranc['Turma_norm'] = df_tranc['Turma'].astype(str).str.strip().str.upper()
    df_tranc = (df_tranc
                .sort_values('Data da solicitação')
                .drop_duplicates(['Ra do Aluno', 'Turma_norm'], keep='last'))

    n_tranc_cursos = len(df_tranc)
    n_tranc_alunos = df_tranc['Ra do Aluno'].nunique()
    print(f"        → {n_tranc_alunos:,} alunos únicos / {n_tranc_cursos:,} combinações aluno+curso")

    # ── Cruzamento ───────────────────────────────────────────
    print("  [4/6] Cruzando bases por RA do aluno...")
    # Montar colunas do requerimento para o merge
    cols_req = ['Ra do Aluno', 'Turma_norm', 'Data da solicitação',
                'Motivo do Requerimento', 'Nome do curso']
    if 'Unidade' in df_tranc.columns:
        cols_req.append('Unidade')
    # Filtrar apenas colunas que existem na base
    cols_req = [c for c in cols_req if c in df_tranc.columns]

    df_tranc_sel = df_tranc[cols_req].rename(
        columns={'Ra do Aluno': 'Ra Aluno'})

    # Se auditoria também tem Unidade, renomear a do requerimento para evitar conflito
    if 'Unidade' in df_audit.columns and 'Unidade' in df_tranc_sel.columns:
        df_tranc_sel = df_tranc_sel.rename(columns={'Unidade': 'Unidade_REQ'})

    # Cruzar por RA + Turma — garante que o trancamento só afeta o curso solicitado
    df_merge = df_audit.merge(
        df_tranc_sel,
        on=['Ra Aluno', 'Turma_norm'],
        how='left'
    )

    matches = df_merge['Data da solicitação'].notna().sum()
    alunos_match = df_merge.loc[df_merge['Data da solicitação'].notna(), 'Ra Aluno'].nunique()
    print(f"        → {matches:,} parcelas vinculadas ao curso trancado ({alunos_match:,} alunos)")

    # ── Classificação ────────────────────────────────────────
    print("  [5/6] Classificando parcelas...")
    resultados = df_merge.apply(
        lambda row: classificar(
            row['tipo_plano'],
            row['Data da solicitação'],
            row['Vencimento'],
            row['foi_pago']
        ), axis=1
    )
    df_merge['CLASSIFICAÇÃO'] = resultados.apply(lambda x: x[0])
    df_merge['MOTIVO']        = resultados.apply(lambda x: x[1])

    # ── Montar tabelas ───────────────────────────────────────
    # Unidade pode ter nome diferente dependendo da base — buscar dinamicamente
    col_unidade = None
    for candidato in ['Unidade', 'unidade', 'UNIDADE', 'Unidade_REQ',
                      'Unidade Ensino', 'Polo', 'polo']:
        if candidato in df_merge.columns:
            col_unidade = candidato
            break
    if col_unidade is None:
        df_merge['_Unidade'] = ''
        col_unidade = '_Unidade'

    # Curso do requerimento (pode ter nome diferente)
    col_curso_req = 'Nome do curso' if 'Nome do curso' in df_merge.columns else None
    if col_curso_req:
        df_merge['Curso Trancado'] = df_merge[col_curso_req].fillna('')
    else:
        df_merge['Curso Trancado'] = df_merge['Curso'].fillna('')

    df_detail = df_merge[[
        'Ra Aluno', 'Aluno', 'CPF Aluno', 'Turma', 'Curso',
        'Curso Trancado', 'Período', 'Status Período', 'Parc. Num.',
        'Vencimento', 'Valor Líquido Previsto', 'tipo_plano',
        'Data da solicitação', 'foi_pago', 'CLASSIFICAÇÃO', 'MOTIVO', col_unidade
    ]].copy()
    df_detail.columns = [
        'RA Aluno', 'Aluno', 'CPF', 'Turma', 'Curso',
        'Curso Trancado', 'Período', 'Status Período', 'Parc. Nº',
        'Vencimento', 'Valor Devido', 'Tipo Plano',
        'Data Solicitação', 'Já Pago?', 'CLASSIFICAÇÃO', 'Motivo', 'Unidade'
    ]
    df_detail['Vencimento'] = (
        df_detail['Vencimento'].dt.strftime('%d/%m/%Y').fillna(''))
    df_detail['Data Solicitação'] = (
        pd.to_datetime(df_detail['Data Solicitação'], errors='coerce')
        .dt.strftime('%d/%m/%Y').fillna(''))
    df_detail['Já Pago?'] = df_detail['Já Pago?'].map(
        {True: 'SIM', False: 'NÃO'})

    # ── Regra: considerar indevidas SOMENTE de quem está na base de requerimentos ──
    # Alunos sem cruzamento (SEM TRANCAMENTO NA BASE) ficam apenas no
    # Detalhamento Completo para referência, mas NÃO entram nos KPIs nem
    # nas abas de Indevidas / Devidas / Manter — evita inflação de valores.
    df_detail_req = df_detail[df_detail['CLASSIFICAÇÃO'] != 'SEM TRANCAMENTO NA BASE'].copy()

    df_indevidas = df_detail_req[df_detail_req['CLASSIFICAÇÃO'] == 'INDEVIDA'].copy()
    df_devidas   = df_detail_req[df_detail_req['CLASSIFICAÇÃO'] == 'DEVIDA'].copy()
    df_manter    = df_detail_req[df_detail_req['CLASSIFICAÇÃO'] == 'MANTER (PAGO)'].copy()  # legado PST

    df_resumo = (
        df_indevidas
        .groupby(['RA Aluno', 'Aluno', 'CPF', 'Tipo Plano',
                  'Status Período', 'Data Solicitação'])
        .agg(
            Qtd=('Parc. Nº', 'count'),
            Total=('Valor Devido', 'sum'),
            Primeira=('Vencimento', 'first'),
            Ultima=('Vencimento', 'last')
        ).reset_index()
    )
    df_resumo.columns = [
        'RA Aluno', 'Aluno', 'CPF', 'Tipo Plano', 'Status Período',
        'Data Solicitação', 'Qtd. Parcelas', 'Valor Total Indevido',
        'Primeira Parcela', 'Última Parcela'
    ]
    df_resumo = df_resumo.sort_values('Valor Total Indevido', ascending=False)

    # KPIs — baseados apenas nos alunos da base de requerimentos
    ti  = df_indevidas['Valor Devido'].sum()       # Total indevido em R$
    aa  = df_indevidas['RA Aluno'].nunique()       # Alunos com parcelas indevidas
    pi  = len(df_indevidas)                        # Qtd parcelas indevidas
    tm  = ti / aa if aa > 0 else 0                 # Ticket médio indevido/aluno
    tma = df_manter['RA Aluno'].nunique()          # Alunos com parcelas já pagas

    td  = df_devidas['Valor Devido'].sum()         # Total devido em R$
    pd_ = len(df_devidas)                          # Qtd parcelas devidas
    ad  = df_devidas['RA Aluno'].nunique()         # Alunos com parcelas devidas
    tg  = ti + td                                  # Total geral auditado em R$
    pg  = pi + pd_                                 # Total geral de parcelas

    # Total auditado = somente quem cruzou com a base de requerimentos
    # (alunos sem requerimento ficam no Detalhamento Completo mas fora dos KPIs)
    total_auditado = len(df_detail_req)

    # ── Gerar planilha ───────────────────────────────────────
    print("  [6/6] Gerando planilha Excel...")
    wb = Workbook()

    # ────────────────────────────────────────────────────────
    # ABA 1 — DASHBOARD
    # ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '📊 Dashboard'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 3
    for col, w in zip(['B', 'C', 'D', 'E', 'F', 'G'],
                      [34, 20, 20, 20, 20, 20]):
        ws1.column_dimensions[col].width = w

    # Título
    ws1.merge_cells('B1:G1')
    c = ws1['B1']
    c.value = '⚠  AUDITORIA FINANCEIRA — TRANCAMENTOS DE MATRÍCULA'
    c.fill  = fill(AZUL_ESC)
    c.font  = font(BRANCO, 14, True)
    c.alignment = aln('center')
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells('B2:G2')
    c = ws1['B2']
    c.value = 'DÉBITOS INDEVIDOS · Análise de Parcelas em Aberto'
    c.fill  = fill(AZUL_MED)
    c.font  = font(BRANCO, 11, False, True)
    c.alignment = aln('center')
    ws1.row_dimensions[2].height = 22

    ws1.merge_cells('B3:G3')
    c = ws1['B3']
    c.value = (f'Base: Requerimentos Deferidos  ·  '
               f'Processado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}  ·  '
               f'Total auditado: {total_auditado:,} registros')
    c.fill  = fill('E8F1FB')
    c.font  = font(AZUL_ESC, 9, False, True)
    c.alignment = aln('center')
    ws1.row_dimensions[3].height = 18
    ws1.row_dimensions[4].height = 8

    def fmt_brl(v):
        return f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    # ── Separador: bloco INDEVIDAS ───────────────────────────
    ws1.merge_cells('B5:G5')
    c = ws1['B5']
    c.value = '🔴  PARCELAS INDEVIDAS'
    c.fill  = fill(VERM)
    c.font  = font(BRANCO, 10, True)
    c.alignment = aln('center')
    ws1.row_dimensions[5].height = 20

    # Labels KPIs — linha INDEVIDAS
    kpi_labels_inv = ['TOTAL INDEVIDO', 'ALUNOS AFETADOS', 'PARCELAS INDEVIDAS',
                      'TICKET MÉDIO/ALUNO', 'MANTER (JÁ PAGOS)', 'TOTAL AUDITADO']
    for col, label in zip(['B', 'C', 'D', 'E', 'F', 'G'], kpi_labels_inv):
        c = ws1[f'{col}6']
        c.value = label
        c.fill  = fill('8B1A1A')
        c.font  = font(BRANCO, 9, True)
        c.alignment = aln('center')
    ws1.row_dimensions[6].height = 18

    # Valores KPIs — linha INDEVIDAS
    kpi_vals_inv = [
        (fmt_brl(ti),           VERM_CLR,  VERM),
        (f'{aa:,}',             VERM_CLR,  VERM),
        (f'{pi:,}',             VERM_CLR,  VERM),
        (fmt_brl(tm),           GOLD_CLR,  '8a5e00'),
        (f'{tma:,}',            GOLD_CLR,  '8a5e00'),
        (f'{total_auditado:,}', 'F4F6F9',  '555555'),
    ]
    for col, (val, bg, fg) in zip(['B', 'C', 'D', 'E', 'F', 'G'], kpi_vals_inv):
        c = ws1[f'{col}7']
        c.value = val
        c.fill  = fill(bg)
        c.font  = Font(name='Calibri', size=18, color=fg, bold=True)
        c.alignment = aln('center')
    ws1.row_dimensions[7].height = 42

    ws1.row_dimensions[8].height = 6

    # ── Separador: bloco DEVIDAS ─────────────────────────────
    ws1.merge_cells('B9:G9')
    c = ws1['B9']
    c.value = '🟢  PARCELAS DEVIDAS'
    c.fill  = fill(VERDE_ESC)
    c.font  = font(BRANCO, 10, True)
    c.alignment = aln('center')
    ws1.row_dimensions[9].height = 20

    # Labels KPIs — linha DEVIDAS
    kpi_labels_dev = ['TOTAL DEVIDO', 'ALUNOS COM DEVIDAS', 'PARCELAS DEVIDAS',
                      '', '', '']
    for col, label in zip(['B', 'C', 'D', 'E', 'F', 'G'], kpi_labels_dev):
        c = ws1[f'{col}10']
        c.value = label
        c.fill  = fill('1B5E20')
        c.font  = font(BRANCO, 9, True)
        c.alignment = aln('center')
    ws1.row_dimensions[10].height = 18

    # Valores KPIs — linha DEVIDAS
    kpi_vals_dev = [
        (fmt_brl(td), VERDE_CLR, VERDE_ESC),
        (f'{ad:,}',   VERDE_CLR, VERDE_ESC),
        (f'{pd_:,}',  VERDE_CLR, VERDE_ESC),
        ('', BRANCO, BRANCO),
        ('', BRANCO, BRANCO),
        ('', BRANCO, BRANCO),
    ]
    for col, (val, bg, fg) in zip(['B', 'C', 'D', 'E', 'F', 'G'], kpi_vals_dev):
        c = ws1[f'{col}11']
        c.value = val
        c.fill  = fill(bg)
        c.font  = Font(name='Calibri', size=18, color=fg, bold=True)
        c.alignment = aln('center')
    ws1.row_dimensions[11].height = 42

    ws1.row_dimensions[12].height = 6

    # ── Separador: bloco TOTAL GERAL ─────────────────────────
    ws1.merge_cells('B13:G13')
    c = ws1['B13']
    c.value = '📊  TOTAL GERAL AUDITADO'
    c.fill  = fill(AZUL_ESC)
    c.font  = font(BRANCO, 10, True)
    c.alignment = aln('center')
    ws1.row_dimensions[13].height = 20

    # Labels KPIs — Total Geral
    kpi_labels_tg = ['VALOR TOTAL AUDITADO', 'TOTAL DE PARCELAS', '', '', '', '']
    for col, label in zip(['B', 'C', 'D', 'E', 'F', 'G'], kpi_labels_tg):
        c = ws1[f'{col}14']
        c.value = label
        c.fill  = fill(AZUL_MED)
        c.font  = font(BRANCO, 9, True)
        c.alignment = aln('center')
    ws1.row_dimensions[14].height = 18

    # Valores KPIs — Total Geral
    kpi_vals_tg = [
        (fmt_brl(tg), AZUL_CLR, AZUL_ESC),
        (f'{pg:,}',   AZUL_CLR, AZUL_ESC),
        ('', BRANCO, BRANCO),
        ('', BRANCO, BRANCO),
        ('', BRANCO, BRANCO),
        ('', BRANCO, BRANCO),
    ]
    for col, (val, bg, fg) in zip(['B', 'C', 'D', 'E', 'F', 'G'], kpi_vals_tg):
        c = ws1[f'{col}15']
        c.value = val
        c.fill  = fill(bg)
        c.font  = Font(name='Calibri', size=18, color=fg, bold=True)
        c.alignment = aln('center')
    ws1.row_dimensions[15].height = 42

    ws1.row_dimensions[16].height = 8

    # Legenda de classificações
    ws1.merge_cells('B17:G17')
    c = ws1['B17']
    c.value = 'LEGENDA DE CLASSIFICAÇÕES'
    c.fill  = fill(AZUL_ESC)
    c.font  = font(BRANCO, 10, True)
    c.alignment = aln('center')
    ws1.row_dimensions[17].height = 22

    legendas = [
        ('INDEVIDA',               VERM_CLR,  VERM,
         'Parcela que deve ser cancelada ou estornada no sistema'),
        ('DEVIDA',                 VERDE_CLR, VERDE_ESC,
         'Parcela legítima — manter a cobrança'),
        ('MANTER (PAGO)',          GOLD_CLR,  '8a5e00',
         'Aluno já pagou — manter; estornar somente se reclamar formalmente'),
        ('VERIFICAR MANUALMENTE',  'F4F6F9',  '555555',
         'Plano não identificado — análise manual necessária'),
        ('SEM TRANCAMENTO NA BASE','F8F8F8',  '888888',
         'Aluno não encontrado na base de requerimentos deferidos'),
    ]
    for i, (status, bg, fg, desc) in enumerate(legendas, 18):
        c = ws1[f'B{i}']
        c.value = status
        c.fill  = fill(bg)
        c.font  = font(fg, 10, True)
        c.alignment = aln('center')
        c.border = brd()
        ws1.merge_cells(f'C{i}:G{i}')
        c2 = ws1[f'C{i}']
        c2.value = desc
        c2.fill  = fill(bg)
        c2.font  = font(fg, 10)
        c2.alignment = aln('left')
        c2.border = brd()
        ws1.row_dimensions[i].height = 22

    ws1.row_dimensions[23].height = 8

    # Distribuição por plano
    ws1.merge_cells('B24:D24')
    c = ws1['B24']
    c.value = 'INDEVIDAS POR PLANO'
    c.fill  = fill(AZUL_ESC)
    c.font  = font(BRANCO, 10, True)
    c.alignment = aln('center')
    ws1.row_dimensions[24].height = 22

    for col, hdr in zip(['B', 'C', 'D'],
                        ['Plano', 'Qtd. Parcelas', 'Valor Indevido']):
        c = ws1[f'{col}25']
        c.value = hdr
        c.fill  = fill(AZUL_MED)
        c.font  = font(BRANCO, 10, True)
        c.alignment = aln('center')
    ws1.row_dimensions[25].height = 20

    plano_group = (df_indevidas
                   .groupby('Tipo Plano')
                   .agg(Qtd=('Parc. Nº', 'count'), Val=('Valor Devido', 'sum'))
                   .reset_index())
    for ri, (_, row) in enumerate(plano_group.iterrows(), 26):
        bg = 'E8F1FB' if ri % 2 == 0 else BRANCO
        for col, val in zip(['B', 'C', 'D'],
                            [row['Tipo Plano'], int(row['Qtd']), fmt_brl(row['Val'])]):
            c = ws1[f'{col}{ri}']
            c.value = val
            c.fill  = fill(bg)
            c.font  = font('1E1E1E', 10)
            c.alignment = aln('center')
            c.border = brd_bottom()
        ws1.row_dimensions[ri].height = 20

    # ────────────────────────────────────────────────────────
    # COLUNAS COMUNS PARA AS ABAS DE DADOS
    # ────────────────────────────────────────────────────────
    COLS_DETAIL = [
        'RA Aluno', 'Aluno', 'CPF', 'Turma', 'Curso',
        'Curso Trancado', 'Período', 'Status Período', 'Parc. Nº',
        'Vencimento', 'Valor Devido', 'Tipo Plano',
        'Data Solicitação', 'Já Pago?', 'CLASSIFICAÇÃO', 'Motivo', 'Unidade'
    ]
    WIDTHS_DETAIL = [12, 30, 15, 14, 30, 30, 10, 14, 10, 13, 14,
                     12, 16, 10, 22, 55, 30]
    CENTRO   = ['RA Aluno', 'Parc. Nº', 'Vencimento', 'Valor Devido',
                'Tipo Plano', 'Data Solicitação', 'Já Pago?', 'CLASSIFICAÇÃO',
                'Status Período']
    MOEDA    = ['Valor Devido']
    WRAP     = ['Motivo']

    def preparar_aba_detalhe(ws, df, titulo, bg_header=AZUL_ESC):
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'

        ws.merge_cells(f'A1:{get_column_letter(len(COLS_DETAIL))}1')
        c = ws['A1']
        c.value = titulo
        c.fill  = fill(bg_header)
        c.font  = font(BRANCO, 12, True)
        c.alignment = aln('center')
        ws.row_dimensions[1].height = 30

        for ci, (col, w) in enumerate(zip(COLS_DETAIL, WIDTHS_DETAIL), 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        df_show = df[COLS_DETAIL].copy()
        last_row = escrever_df(ws, df_show, start_row=2,
                               header_bg=bg_header,
                               colunas_centro=CENTRO,
                               colunas_moeda=MOEDA,
                               colunas_wrap=WRAP)

        # Linha de total — encontrar coluna de Valor Devido dinamicamente
        total_row = last_row + 1
        n_cols = len(COLS_DETAIL)

        # Encontrar índice da coluna Valor Devido (base 1)
        try:
            col_valor_idx = COLS_DETAIL.index('Valor Devido') + 1
        except ValueError:
            col_valor_idx = 10  # fallback

        # Merge da coluna A até a coluna anterior ao Valor Devido
        merge_ate = max(col_valor_idx - 1, 1)
        if merge_ate > 1:
            ws.merge_cells(
                f'A{total_row}:{get_column_letter(merge_ate)}{total_row}')

        c = ws[f'A{total_row}']
        c.value = f'TOTAL — {len(df_show):,} registros'
        c.fill  = fill(AZUL_ESC)
        c.font  = font(BRANCO, 10, True)
        c.alignment = aln('right')

        # Total valor na coluna correta
        total_val = df_show['Valor Devido'].sum() if 'Valor Devido' in df_show.columns else 0
        c2 = ws.cell(row=total_row, column=col_valor_idx)
        c2.value = total_val
        c2.number_format = 'R$ #,##0.00'
        c2.fill = fill(VERDE_CLR)
        c2.font = font(VERDE_ESC, 11, True)
        c2.alignment = aln('center')
        ws.row_dimensions[total_row].height = 24

    # ────────────────────────────────────────────────────────
    # ABA 2 — PARCELAS INDEVIDAS
    # ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('📋 Parcelas Indevidas')
    preparar_aba_detalhe(
        ws2, df_indevidas,
        '📋  PARCELAS INDEVIDAS — DETALHAMENTO COMPLETO',
        bg_header=AZUL_ESC
    )

    # ────────────────────────────────────────────────────────
    # ABA 3 — RESUMO POR ALUNO
    # ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('👤 Resumo por Aluno')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A3'

    ws3.merge_cells('A1:J1')
    c = ws3['A1']
    c.value = '👤  RESUMO POR ALUNO — VALOR TOTAL INDEVIDO'
    c.fill  = fill(AZUL_ESC)
    c.font  = font(BRANCO, 12, True)
    c.alignment = aln('center')
    ws3.row_dimensions[1].height = 30

    cols3 = ['RA Aluno', 'Aluno', 'CPF', 'Tipo Plano', 'Status Período',
             'Data Solicitação', 'Qtd. Parcelas', 'Valor Total Indevido',
             'Primeira Parcela', 'Última Parcela']
    widths3 = [12, 32, 15, 12, 14, 16, 14, 20, 16, 16]

    for ci, (col, w) in enumerate(zip(cols3, widths3), 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
        c = ws3.cell(row=2, column=ci)
        c.value = col
        c.fill  = fill(AZUL_ESC)
        c.font  = font(BRANCO, 10, True)
        c.alignment = aln('center')
        c.border = brd()
    ws3.row_dimensions[2].height = 22

    for ri, (_, row) in enumerate(df_resumo.iterrows(), 3):
        bg = AZUL_CLR if (ri - 3) % 2 == 0 else BRANCO
        ws3.row_dimensions[ri].height = 16
        for ci, col in enumerate(cols3, 1):
            c = ws3.cell(row=ri, column=ci)
            val = row[col]
            if col == 'Valor Total Indevido':
                try:
                    c.value = float(val)
                    c.number_format = 'R$ #,##0.00'
                except Exception:
                    c.value = val
                c.font = font(VERM, 10, True)
            else:
                c.value = val if not (isinstance(val, float) and
                                      pd.isna(val)) else ''
                c.font = font('1E1E1E', 9)
            c.fill = fill(bg)
            c.alignment = aln('center' if ci in [1, 4, 5, 6, 7, 8, 9, 10]
                               else 'left')
            c.border = brd_bottom()

    last3 = 2 + len(df_resumo) + 1
    # Coluna de Valor Total Indevido é a 8ª coluna no resumo
    col_val_resumo = cols3.index('Valor Total Indevido') + 1
    merge_ate3 = max(col_val_resumo - 1, 1)
    if merge_ate3 > 1:
        ws3.merge_cells(f'A{last3}:{get_column_letter(merge_ate3)}{last3}')
    c = ws3[f'A{last3}']
    c.value = f'TOTAL — {len(df_resumo):,} alunos com parcelas indevidas'
    c.fill  = fill(AZUL_ESC)
    c.font  = font(BRANCO, 10, True)
    c.alignment = aln('right')
    c2 = ws3.cell(row=last3, column=col_val_resumo)
    c2.value = df_resumo['Valor Total Indevido'].sum()
    c2.number_format = 'R$ #,##0.00'
    c2.fill = fill(VERDE_CLR)
    c2.font = font(VERDE_ESC, 11, True)
    c2.alignment = aln('center')
    ws3.row_dimensions[last3].height = 24

    # ────────────────────────────────────────────────────────
    # ABA 4 — MANTER (JÁ PAGOS)
    # ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet('✅ Manter (Já Pagos)')
    preparar_aba_detalhe(
        ws4, df_manter,
        '✅  MANTER (JÁ PAGOS) — Estornar somente se o aluno reclamar formalmente',
        bg_header=VERDE_ESC
    )

    # ────────────────────────────────────────────────────────
    # ABA 5 — DETALHAMENTO COMPLETO
    # ────────────────────────────────────────────────────────
    ws5 = wb.create_sheet('📑 Detalhamento Completo')
    preparar_aba_detalhe(
        ws5, df_detail,
        '📑  DETALHAMENTO COMPLETO — TODOS OS REGISTROS COM CLASSIFICAÇÃO',
        bg_header=AZUL_MED
    )

    return wb, ti, aa, pi, tm, tma, total_auditado, td, ad, pd_, tg, pg

# ════════════════════════════════════════════════════════════
# EXECUÇÃO PRINCIPAL
# ════════════════════════════════════════════════════════════
def main():
    limpar()
    cabecalho()

    print("📂 PASSO 1 — BASE DE AUDITORIA FINANCEIRA")
    print("-" * 60)
    caminho_audit = pedir_arquivo(
        "Informe o caminho completo do arquivo de Auditoria Financeira:")

    print("📂 PASSO 2 — BASE DE REQUERIMENTOS DEFERIDOS")
    print("-" * 60)
    caminho_req = pedir_arquivo(
        "Informe o caminho completo do arquivo de Requerimentos:")

    print()
    print("⚙  PROCESSANDO...")
    print("-" * 60)

    try:
        wb, ti, aa, pi, tm, tma, total, td, ad, pd_, tg, pg = processar(caminho_audit, caminho_req)
    except KeyError as e:
        print(f"\n❌ ERRO: Coluna não encontrada no arquivo: {e}")
        print("   Verifique se você selecionou os arquivos corretos.")
        input("\nPressione Enter para fechar...")
        return
    except Exception as e:
        print(f"\n❌ ERRO inesperado: {e}")
        input("\nPressione Enter para fechar...")
        return

    # Salvar na mesma pasta do arquivo de auditoria
    pasta = os.path.dirname(os.path.abspath(caminho_audit))
    hoje  = datetime.now().strftime('%d%m%Y_%H%M')
    nome_saida = f'Cruzamento_Trancamentos_{hoje}.xlsx'
    caminho_saida = os.path.join(pasta, nome_saida)
    wb.save(caminho_saida)

    # Resultado final
    limpar()
    cabecalho()
    print("✅  PROCESSAMENTO CONCLUÍDO COM SUCESSO!")
    print("=" * 60)
    print()
    print("  🔴  PARCELAS INDEVIDAS")
    print(f"  📊 Total Indevido:      R$ {ti:,.2f}".replace(',','X').replace('.',',').replace('X','.'))
    print(f"  👥 Alunos Afetados:     {aa:,}")
    print(f"  📋 Parcelas Indevidas:  {pi:,}")
    print(f"  💰 Ticket Médio/Aluno:  R$ {tm:,.2f}".replace(',','X').replace('.',',').replace('X','.'))
    print(f"  ✅ Manter (Já Pagos):   {tma:,} alunos")
    print()
    print("  🟢  PARCELAS DEVIDAS")
    print(f"  📊 Total Devido:        R$ {td:,.2f}".replace(',','X').replace('.',',').replace('X','.'))
    print(f"  👥 Alunos c/ Devidas:   {ad:,}")
    print(f"  📋 Parcelas Devidas:    {pd_:,}")
    print()
    print("  📊  TOTAL GERAL AUDITADO")
    print(f"  💵 Valor Total:         R$ {tg:,.2f}".replace(',','X').replace('.',',').replace('X','.'))
    print(f"  📋 Total de Parcelas:   {pg:,}")
    print(f"  📑 Registros Auditados: {total:,}")
    print()
    print("=" * 60)
    print(f"  📁 Arquivo salvo em:")
    print(f"     {caminho_saida}")
    print("=" * 60)
    print()
    input("  Pressione Enter para fechar...")

if __name__ == '__main__':
    main()
